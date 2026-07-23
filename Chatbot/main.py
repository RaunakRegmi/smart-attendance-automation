import pandas as pd
import sqlite3
import os
import re
from datetime import datetime

EXCEL_FILE = "Attendance Format.xlsx"
DB_FILE = "attendance.db"
OUTPUT_DIR = "output"

os.makedirs(OUTPUT_DIR, exist_ok=True)


def normalize_label(value):
    text = str(value).strip()
    text = re.sub(r"[^0-9a-zA-Z]+", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    return text.lower()


def validate_required_columns(df, required_columns, context):
    missing = [column for column in required_columns if column not in df.columns]
    if missing:
        raise ValueError(f"{context} is missing required columns: {', '.join(missing)}")


def get_week_slot_columns(df):
    week_pattern = re.compile(r"^week_(\d+)_([a-z]+)$")
    week_slots = {}

    for column in df.columns:
        match = week_pattern.match(column)
        if not match:
            continue

        week_number = int(match.group(1))
        week_slots.setdefault(week_number, []).append(column)

    for week_number in week_slots:
        week_slots[week_number].sort()

    return dict(sorted(week_slots.items()))


def load_attendance_sheet(filepath):
    raw = pd.read_excel(filepath, sheet_name="Attendance Sheet", header=None)

    header_row = raw.iloc[4]
    sub_row = raw.iloc[5]
    date_row = raw.iloc[6]

    col_names = []
    current_week = None
    for i, (h, s) in enumerate(zip(header_row, sub_row)):
        if pd.notna(h) and str(h).strip() not in ("", "nan"):
            current_week = str(h).strip()
        header_text = str(h).strip() if pd.notna(h) else ""
        sub_text = str(s).strip() if pd.notna(s) else ""

        if i == 0:
            col_names.append("sno")
        elif i == 1:
            col_names.append("batch")
        elif i == 2:
            col_names.append("student_name")
        elif i == 3:
            col_names.append("email")
        elif header_text.lower().startswith("total "):
            col_names.append(normalize_label(header_text))
        elif header_text and sub_text:
            week_label = normalize_label(current_week) if current_week else f"week_{i}"
            col_names.append(f"{week_label}_{normalize_label(sub_text)}")
        elif header_text:
            col_names.append(normalize_label(header_text))
        else:
            col_names.append(f"col_{i}")

    col_names = col_names[:len(raw.columns)]
    if len(col_names) < len(raw.columns):
        col_names = col_names + [f"extra_{i}" for i in range(len(raw.columns) - len(col_names))]
    raw.columns = col_names
    validate_required_columns(
        raw,
        ["sno", "batch", "student_name", "email", "total_present", "total_absent", "total_late", "total_present_with_late", "total_classes"],
        "Attendance Sheet",
    )

    df = raw.iloc[7:].copy()
    df = df.reset_index(drop=True)

    # Keep only rows where S.NO is numeric (drop totals row at bottom)
    df = df[pd.to_numeric(df["sno"], errors="coerce").notna()]
    df["sno"] = pd.to_numeric(df["sno"], errors="coerce").astype("Int64")

    return df, date_row


def extract_student_summary(df):
    """Pull the clean summary columns."""
    sno_col = "sno"
    batch_col = "batch"
    name_col = "student_name"
    email_col = "email"
    total_present_col = "total_present"
    total_absent_col = "total_absent"
    total_late_col = "total_late"
    total_present_with_late_col = "total_present_with_late"
    total_classes_col = "total_classes"
    percentage_col = "total_present_in_percentage"

    students = []
    for _, row in df.iterrows():
        sno = row[sno_col]
        batch = row.get(batch_col, "")
        name = row.get(name_col, f"Student_{int(float(sno))}")
        email = row.get(email_col, "")

        def safe_int(col):
            if col is None:
                return 0
            v = row.get(col, 0)
            try:
                return int(float(v)) if pd.notna(v) else 0
            except (ValueError, TypeError):
                return 0

        def safe_float(col):
            if col is None:
                return 0.0
            v = row.get(col, 0)
            try:
                return round(float(v), 2) if pd.notna(v) else 0.0
            except (ValueError, TypeError):
                return 0.0

        total_present = safe_int(total_present_col)
        total_absent = safe_int(total_absent_col)
        total_late = safe_int(total_late_col)
        total_present_with_late = safe_int(total_present_with_late_col)
        total_classes = safe_int(total_classes_col)
        percentage = safe_float(percentage_col)

        if percentage == 0.0 and total_classes:
            percentage = round((total_present_with_late / total_classes) * 100, 2)

        if total_classes == 0 and total_present == 0:
            continue

        students.append({
            "sno": int(sno),
            "batch": str(batch).strip(),
            "student_name": str(name).strip(),
            "email": str(email).strip(),
            "total_present": total_present,
            "total_absent": total_absent,
            "total_late": total_late,
            "total_present_with_late": total_present_with_late,
            "total_classes": total_classes,
            "attendance_percentage": percentage,
            "status": classify_status(percentage),
        })

    return pd.DataFrame(students)


def classify_status(pct):
    if pct >= 90:
        return "Excellent"
    elif pct >= 75:
        return "Satisfactory"
    elif pct >= 60:
        return "At Risk"
    else:
        return "Critical"


def load_weekly_report(filepath):
    raw = pd.read_excel(filepath, sheet_name="Weekly Report", header=None)

    section_row = raw.iloc[0].tolist()
    header_row = raw.iloc[1].tolist()
    cleaned_headers = []
    current_section = None

    for i, (section, header) in enumerate(zip(section_row, header_row)):
        if pd.notna(section) and str(section).strip():
            current_section = str(section).strip()

        section_label = normalize_label(current_section) if current_section else "report"
        header_label = normalize_label(header) if pd.notna(header) and str(header).strip() else ""

        if i == 0:
            cleaned_headers.append("sno")
        elif i == 1:
            cleaned_headers.append("student_name")
        elif header_label:
            if 7 <= i <= 15 and current_section:
                cleaned_headers.append(f"{section_label}_{header_label}")
            else:
                cleaned_headers.append(header_label)
        elif section_label:
            cleaned_headers.append(section_label)
        else:
            cleaned_headers.append(f"col_{i}")

    df = raw.iloc[2:].copy()
    df.columns = cleaned_headers
    validate_required_columns(df, ["sno"], "Weekly Report")
    df = df[pd.to_numeric(df["sno"], errors="coerce").notna()]
    df = df.reset_index(drop=True)
    df["sno"] = pd.to_numeric(df["sno"], errors="coerce").astype("Int64")
    return df


def build_student_report(df_attendance, df_weekly):
    attendance_cols = [
        "sno",
        "student_name",
        "batch",
        "email",
        "total_present",
        "total_absent",
        "total_late",
        "total_present_with_late",
        "total_classes",
        "attendance_percentage",
        "status",
    ]

    attendance_lookup = df_attendance[attendance_cols].copy()
    attendance_lookup["sno"] = attendance_lookup["sno"].astype("Int64")

    report = df_weekly.copy()
    if "student_name" in report.columns:
        report = report.rename(columns={"student_name": "weekly_student_name"})
    report["sno"] = report["sno"].astype("Int64")
    report = report.merge(attendance_lookup, on="sno", how="left")

    if "weekly_student_name" in report.columns:
        report["student_name"] = report["student_name"].fillna(report["weekly_student_name"])
        report = report.drop(columns=["weekly_student_name"])

    return report


def store_to_sqlite(df_attendance, df_student_report, db_path, analytics_dict):
    """Store all dataframes to SQLite with created_at timestamp."""
    now = datetime.now().isoformat()
    
    with sqlite3.connect(db_path) as conn:
        # Core tables
        attendance_to_store = df_attendance.copy()
        attendance_to_store["created_at"] = now
        attendance_to_store.to_sql("attendance_summary", conn, if_exists="replace", index=False)
        
        student_report_to_store = df_student_report.copy()
        student_report_to_store["created_at"] = now
        student_report_to_store.to_sql("student_reports", conn, if_exists="replace", index=False)
        
        for table_name, df in analytics_dict.items():
            if not df.empty:
                df_to_store = df.copy()
                df_to_store["created_at"] = now
                df_to_store.to_sql(table_name, conn, if_exists="replace", index=False)

    print(f"  ✓ SQLite DB saved: {db_path}")
    print(f"    - {len(df_attendance)} attendance records")
    print(f"    - {len(df_student_report)} student reports")
    for table_name, df in analytics_dict.items():
        if not df.empty:
            print(f"    - {len(df)} {table_name} records")


def export_csvs(df, output_dir):
    full_path = os.path.join(output_dir, "attendance_summary.csv")
    df.to_csv(full_path, index=False)
    print(f"  ✓ Full summary CSV: {full_path}")

    at_risk = df[df["attendance_percentage"] < 75].copy()
    at_risk_path = os.path.join(output_dir, "at_risk_students.csv")
    at_risk.to_csv(at_risk_path, index=False)
    print(f"  ✓ At-risk students CSV: {at_risk_path} ({len(at_risk)} students)")

    batch_summary = df.groupby("batch").agg(
        total_students=("sno", "count"),
        avg_attendance=("attendance_percentage", "mean"),
        below_75=("attendance_percentage", lambda x: (x < 75).sum()),
        perfect_attendance=("attendance_percentage", lambda x: (x == 100).sum()),
    ).reset_index()
    batch_summary["avg_attendance"] = batch_summary["avg_attendance"].round(2)
    batch_path = os.path.join(output_dir, "batch_summary.csv")
    batch_summary.to_csv(batch_path, index=False)
    print(f"  ✓ Batch summary CSV: {batch_path}")


def export_student_report_csv(df_student_report, output_dir):
    report_path = os.path.join(output_dir, "student_reports.csv")
    df_student_report.to_csv(report_path, index=False)
    print(f"  ✓ Student reports CSV: {report_path}")


def export_analytics_csvs(analytics_dict, output_dir):
    """Export all analytics dataframes to CSV."""
    csv_mapping = {
        "engagement_scores": "engagement_scores.csv",
        "attendance_trends": "attendance_trends.csv",
        "batch_percentiles": "student_rankings.csv",
        "course_analytics": "course_performance_summary.csv",
        "risk_assessment": "risk_assessment.csv",
        "batch_analytics": "batch_analytics.csv",
        "demographic_breakdown": "demographic_breakdown.csv",
        "weekly_breakdown": "weekly_attendance_breakdown.csv",
    }
    
    for df_name, csv_name in csv_mapping.items():
        if df_name in analytics_dict and not analytics_dict[df_name].empty:
            path = os.path.join(output_dir, csv_name)
            analytics_dict[df_name].to_csv(path, index=False)
            print(f"  ✓ {csv_name}: {len(analytics_dict[df_name])} rows")



def extract_weekly_breakdown(df_raw):
    """Extract week-by-week attendance detail for each student."""
    weekly_data = []
    week_slots = get_week_slot_columns(df_raw)

    for _, row in df_raw.iterrows():
        sno = int(row["sno"])
        student_name = row.get("student_name", f"Student_{sno}")
        batch = row.get("batch", "")
        email = row.get("email", "")

        week_pattern = {}
        for week_number, columns in week_slots.items():
            values = [row.get(column) for column in columns if pd.notna(row.get(column))]
            present_count = sum(1 for value in values if str(value).strip().lower() == "present")
            week_pattern[week_number] = round(present_count / len(columns), 3) if columns else None

        weekly_data.append({
            "sno": sno,
            "student_name": student_name,
            "batch": batch,
            "email": email,
            **{f"week_{i}_attendance_ratio": v for i, v in week_pattern.items()}
        })

    return pd.DataFrame(weekly_data)


def calculate_engagement_score(df_attendance):
    """Calculate multi-factor engagement score: attendance + consistency + performance."""
    scores = []
    for _, row in df_attendance.iterrows():
        pct = row.get("attendance_percentage", 0)
        total = row.get("total_classes", 1)
        late = row.get("total_late", 0)

        consistency_penalty = 0
        if late > (total * 0.1):
            consistency_penalty += 5
        if late > (total * 0.2):
            consistency_penalty += 10

        attendance_score = pct
        reliability_score = max(0, 100 - (late / total * 100) * 0.5) if total else 100
        engagement_score = (attendance_score * 0.6) + (reliability_score * 0.4) - consistency_penalty
        engagement_score = max(0, min(100, engagement_score))
        engagement_tier = (
            "Highly Engaged" if engagement_score >= 85
            else "Engaged" if engagement_score >= 70
            else "Moderately Engaged" if engagement_score >= 55
            else "Low Engagement" if engagement_score >= 40
            else "Disengaged"
        )

        scores.append({
            "sno": row["sno"],
            "student_name": row["student_name"],
            "batch": row["batch"],
            "engagement_score": round(engagement_score, 2),
            "engagement_tier": engagement_tier,
            "reliability_score": round(reliability_score, 2),
            "consistency_penalty": consistency_penalty,
        })

    return pd.DataFrame(scores)


def calculate_attendance_trends(df_raw):
    """Detect improving vs declining trends by comparing first half vs second half attendance."""
    trends = []
    week_slots = get_week_slot_columns(df_raw)
    first_half_weeks = {week_number for week_number in week_slots if week_number <= 6}
    second_half_weeks = {week_number for week_number in week_slots if week_number >= 7}

    for _, row in df_raw.iterrows():
        sno = int(row["sno"])
        student_name = row.get("student_name", f"Student_{sno}")
        batch = row.get("batch", "")

        first_half_present = 0
        first_half_total = 0
        second_half_present = 0
        second_half_total = 0

        for week_number, columns in week_slots.items():
            values = [row.get(column) for column in columns if pd.notna(row.get(column))]
            present_count = sum(1 for value in values if str(value).strip().lower() == "present")

            if week_number in first_half_weeks:
                first_half_present += present_count
                first_half_total += len(columns)
            elif week_number in second_half_weeks:
                second_half_present += present_count
                second_half_total += len(columns)

        first_half_pct = (first_half_present / first_half_total * 100) if first_half_total else 0
        second_half_pct = (second_half_present / second_half_total * 100) if second_half_total else 0
        trend_change = second_half_pct - first_half_pct

        if trend_change > 5:
            trend = "Improving"
        elif trend_change < -5:
            trend = "Declining"
        else:
            trend = "Stable"

        trends.append({
            "sno": sno,
            "student_name": student_name,
            "batch": batch,
            "first_half_attendance_pct": round(first_half_pct, 2),
            "second_half_attendance_pct": round(second_half_pct, 2),
            "trend_change_pct": round(trend_change, 2),
            "trend_direction": trend,
        })

    return pd.DataFrame(trends)


def calculate_batch_percentiles(df_attendance):
    """Calculate percentile rankings within each batch."""
    result = df_attendance.copy()
    for batch in result["batch"].unique():
        batch_mask = result["batch"] == batch
        result.loc[batch_mask, "percentile_attendance"] = (
            result.loc[batch_mask, "attendance_percentage"].rank(pct=True) * 100
        ).round(2)
        result.loc[batch_mask, "rank_in_batch"] = result.loc[batch_mask, "attendance_percentage"].rank(
            method="min", ascending=False
        ).astype(int)

    return result[["sno", "student_name", "batch", "attendance_percentage", "percentile_attendance", "rank_in_batch"]]


def generate_course_performance_summary(df_weekly):
    """Extract course-level performance analytics from weekly report."""
    if df_weekly.empty:
        return pd.DataFrame()

    courses = {}
    course_cols = {
        "cmp5367_artificial_intelligence_and_machine_learning": ["cmp5367_artificial_intelligence_and_machine_learning_total_classes",
                                                                  "cmp5367_artificial_intelligence_and_machine_learning_total_present",
                                                                  "cmp5367_artificial_intelligence_and_machine_learning_present_percentage"],
        "dig5127_database_web_application_development": ["dig5127_database_web_application_development_total_classes",
                                                         "dig5127_database_web_application_development_total_present",
                                                         "dig5127_database_web_application_development_present_percentage"],
        "cmp5332_object_oriented_programming": ["cmp5332_object_oriented_programming_total_classes",
                                                "cmp5332_object_oriented_programming_total_present",
                                                "cmp5332_object_oriented_programming_present_percentage"],
    }

    course_analytics = []
    for course_name, cols in course_cols.items():
        if all(c in df_weekly.columns for c in cols):
            df_course = df_weekly[cols].copy()
            df_course = df_course.apply(pd.to_numeric, errors="coerce")

            total_classes = df_course[cols[0]].sum()
            total_present = df_course[cols[1]].sum()
            avg_pct = df_course[cols[2]].mean()
            students_100 = (df_course[cols[2]] == 100).sum()
            students_below_75 = (df_course[cols[2]] < 75).sum()

            course_analytics.append({
                "course_name": course_name.replace("_", " ").title(),
                "total_students": len(df_course),
                "total_classes": int(total_classes),
                "total_present": int(total_present),
                "avg_attendance_pct": round(avg_pct, 2),
                "students_perfect": students_100,
                "students_below_75": students_below_75,
                "performance_tier": (
                    "Excellent" if avg_pct >= 90 else
                    "Good" if avg_pct >= 75 else
                    "Average" if avg_pct >= 60 else "Poor"
                ),
            })

    return pd.DataFrame(course_analytics)


def generate_risk_assessment(df_attendance, df_engagement, df_trends):
    """Generate comprehensive risk assessment with multiple factors."""
    risk_data = []
    engagement_lookup = df_engagement.set_index("sno").to_dict("index") if not df_engagement.empty else {}
    trend_lookup = df_trends.set_index("sno").to_dict("index") if not df_trends.empty else {}

    for _, att_row in df_attendance.iterrows():
        sno = att_row["sno"]
        eng_row = engagement_lookup.get(sno)
        trend_row = trend_lookup.get(sno)

        pct = att_row.get("attendance_percentage", 0)
        engagement_score = eng_row.get("engagement_score", 50) if eng_row is not None else 50
        trend = trend_row.get("trend_direction", "Stable") if trend_row is not None else "Stable"

        risk_factors = []
        risk_score = 0

        if pct < 75:
            risk_factors.append("low_attendance")
            risk_score += 40
        elif pct < 85:
            risk_score += 15

        if engagement_score < 60:
            risk_factors.append("low_engagement")
            risk_score += 30

        if trend == "Declining":
            risk_factors.append("declining_trend")
            risk_score += 25

        if att_row.get("total_late", 0) > att_row.get("total_classes", 1) * 0.15:
            risk_factors.append("high_tardiness")
            risk_score += 15

        risk_score = min(100, risk_score)
        risk_level = (
            "Critical" if risk_score >= 75 else
            "High" if risk_score >= 50 else
            "Medium" if risk_score >= 30 else
            "Low"
        )
        intervention_needed = risk_score >= 50

        risk_data.append({
            "sno": sno,
            "student_name": att_row["student_name"],
            "batch": att_row["batch"],
            "attendance_percentage": pct,
            "engagement_score": engagement_score,
            "trend": trend,
            "risk_score": round(risk_score, 2),
            "risk_level": risk_level,
            "risk_factors": "|".join(risk_factors) if risk_factors else "none",
            "intervention_needed": intervention_needed,
        })

    return pd.DataFrame(risk_data)


def generate_batch_analytics(df_attendance):
    """Generate comprehensive batch-level analytics."""
    batch_analytics = []
    for batch in df_attendance["batch"].unique():
        batch_df = df_attendance[df_attendance["batch"] == batch]
        total = len(batch_df)
        avg_pct = batch_df["attendance_percentage"].mean()
        median_pct = batch_df["attendance_percentage"].median()
        std_dev = batch_df["attendance_percentage"].std()

        excellent = (batch_df["attendance_percentage"] >= 90).sum()
        satisfactory = ((batch_df["attendance_percentage"] >= 75) & (batch_df["attendance_percentage"] < 90)).sum()
        at_risk = ((batch_df["attendance_percentage"] >= 60) & (batch_df["attendance_percentage"] < 75)).sum()
        critical = (batch_df["attendance_percentage"] < 60).sum()

        health_score = (excellent * 4 + satisfactory * 3 + at_risk * 2 + critical * 1) / (total * 4) * 100 if total else 0

        batch_analytics.append({
            "batch": batch,
            "total_students": total,
            "avg_attendance_pct": round(avg_pct, 2),
            "median_attendance_pct": round(median_pct, 2),
            "std_deviation": round(std_dev, 2),
            "excellent_count": excellent,
            "satisfactory_count": satisfactory,
            "at_risk_count": at_risk,
            "critical_count": critical,
            "batch_health_score": round(health_score, 2),
            "batch_health_tier": "Excellent" if health_score >= 80 else "Good" if health_score >= 60 else "Fair" if health_score >= 40 else "Poor",
        })

    return pd.DataFrame(batch_analytics)


def generate_demographic_breakdown(df_attendance):
    """Generate detailed demographic and status breakdown."""
    demographic = []
    for batch in df_attendance["batch"].unique():
        for status in ["Excellent", "Satisfactory", "At Risk", "Critical"]:
            count = len(df_attendance[(df_attendance["batch"] == batch) & (df_attendance["status"] == status)])
            demographic.append({"batch": batch, "status": status, "count": count})

    return pd.DataFrame(demographic)


def generate_report_excel(df, output_dir):
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  GradientFill)
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import ColorScaleRule, CellIsRule

    wb = Workbook()

    DARK_BLUE   = "1A3A5C"
    MED_BLUE    = "2E6DA4"
    LIGHT_BLUE  = "D6E8FA"
    ACCENT_TEAL = "0D7680"
    WHITE       = "FFFFFF"
    LIGHT_GREY  = "F5F7FA"
    MID_GREY    = "D0D8E4"
    GREEN       = "27AE60"
    ORANGE      = "E67E22"
    RED         = "E74C3C"
    YELLOW_BG   = "FFF9C4"

    thin = Side(style="thin", color=MID_GREY)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def header_font(size=11, bold=True, color=WHITE):
        return Font(name="Arial", size=size, bold=bold, color=color)

    def body_font(size=10, bold=False, color="222222"):
        return Font(name="Arial", size=size, bold=bold, color=color)

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def center():
        return Alignment(horizontal="center", vertical="center", wrap_text=True)

    def left():
        return Alignment(horizontal="left", vertical="center")

    ws1 = wb.active
    ws1.title = "Attendance Summary"
    ws1.sheet_view.showGridLines = False

    # Title banner
    ws1.merge_cells("A1:J1")
    ws1["A1"] = "🎓  SMART CAMPUS MANAGEMENT SYSTEM — Attendance Report"
    ws1["A1"].font = Font(name="Arial", size=14, bold=True, color=WHITE)
    ws1["A1"].fill = fill(DARK_BLUE)
    ws1["A1"].alignment = center()
    ws1.row_dimensions[1].height = 36

    ws1.merge_cells("A2:J2")
    ws1["A2"] = f"Generated: {datetime.now().strftime('%B %d, %Y')}   |   Batch: {df['batch'].iloc[0] if len(df) else 'N/A'}   |   Total Students: {len(df)}"
    ws1["A2"].font = Font(name="Arial", size=9, color=MED_BLUE)
    ws1["A2"].fill = fill(LIGHT_BLUE)
    ws1["A2"].alignment = center()
    ws1.row_dimensions[2].height = 18

    headers = ["S.No", "Student Name", "Batch", "Email",
               "Total Classes", "Present", "Absent", "Late",
               "Attendance %", "Status"]
    col_widths = [6, 26, 14, 32, 14, 10, 10, 8, 14, 14]

    ws1.row_dimensions[3].height = 22
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws1.cell(row=3, column=col_idx, value=h)
        cell.font = header_font(10)
        cell.fill = fill(MED_BLUE)
        cell.alignment = center()
        cell.border = border
        ws1.column_dimensions[get_column_letter(col_idx)].width = w

    # Data rows
    status_colors = {
        "Excellent":    (GREEN,   "E8F8F0"),
        "Satisfactory": (MED_BLUE, LIGHT_BLUE),
        "At Risk":      (ORANGE,   "FEF3E2"),
        "Critical":     (RED,      "FDEDEC"),
    }

    for row_idx, row in df.iterrows():
        excel_row = row_idx + 4
        bg = LIGHT_GREY if row_idx % 2 == 0 else WHITE
        ws1.row_dimensions[excel_row].height = 18

        values = [
            row["sno"], row["student_name"], row["batch"], row["email"],
            row["total_classes"], row["total_present"], row["total_absent"],
            row["total_late"], row["attendance_percentage"], row["status"]
        ]
        aligns = [center, left, center, left,
                  center, center, center, center, center, center]

        for col_idx, (val, aln) in enumerate(zip(values, aligns), start=1):
            cell = ws1.cell(row=excel_row, column=col_idx, value=val)
            cell.font = body_font()
            cell.alignment = aln()
            cell.border = border

            if col_idx == 9:  # Attendance %
                cell.number_format = "0.00%"
                pct_val = float(val) / 100 if val else 0
                cell.value = pct_val
            elif col_idx == 10:  # Status
                sc, bg_s = status_colors.get(str(val), (MED_BLUE, LIGHT_BLUE))
                cell.font = Font(name="Arial", size=10, bold=True, color=sc)
                cell.fill = fill(bg_s)
            else:
                cell.fill = fill(bg)

    ws1.freeze_panes = "A4"

    ws2 = wb.create_sheet("At-Risk Students")
    ws2.sheet_view.showGridLines = False

    ws2.merge_cells("A1:G1")
    ws2["A1"] = "AT-RISK STUDENTS — Attendance Below 75%"
    ws2["A1"].font = Font(name="Arial", size=13, bold=True, color=WHITE)
    ws2["A1"].fill = fill("B03A2E")
    ws2["A1"].alignment = center()
    ws2.row_dimensions[1].height = 32

    at_risk = df[df["attendance_percentage"] < 75].copy()

    ws2.merge_cells("A2:G2")
    ws2["A2"] = f"Total students requiring attention: {len(at_risk)}"
    ws2["A2"].font = Font(name="Arial", size=10, color="B03A2E", bold=True)
    ws2["A2"].fill = fill("FDEDEC")
    ws2["A2"].alignment = center()
    ws2.row_dimensions[2].height = 18

    ar_headers = ["S.No", "Student Name", "Batch", "Total Classes",
                  "Present", "Absent", "Attendance %"]
    ar_widths = [6, 28, 16, 14, 10, 10, 14]

    ws2.row_dimensions[3].height = 22
    for c, (h, w) in enumerate(zip(ar_headers, ar_widths), 1):
        cell = ws2.cell(row=3, column=c, value=h)
        cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
        cell.fill = fill("C0392B")
        cell.alignment = center()
        cell.border = border
        ws2.column_dimensions[get_column_letter(c)].width = w

    for r_idx, (_, row) in enumerate(at_risk.iterrows()):
        er = r_idx + 4
        bg = "FEF9F9" if r_idx % 2 == 0 else WHITE
        ws2.row_dimensions[er].height = 18
        vals = [row["sno"], row["student_name"], row["batch"],
                row["total_classes"], row["total_present"],
                row["total_absent"], row["attendance_percentage"] / 100]
        alns = [center, left, center, center, center, center, center]
        fmts = [None, None, None, None, None, None, "0.00%"]
        for c, (v, a, f) in enumerate(zip(vals, alns, fmts), 1):
            cell = ws2.cell(row=er, column=c, value=v)
            cell.font = body_font()
            cell.fill = fill(bg)
            cell.alignment = a()
            cell.border = border
            if f:
                cell.number_format = f

    ws2.freeze_panes = "A4"

    ws3 = wb.create_sheet("Statistics")
    ws3.sheet_view.showGridLines = False

    ws3.merge_cells("A1:F1")
    ws3["A1"] = "ATTENDANCE STATISTICS DASHBOARD"
    ws3["A1"].font = Font(name="Arial", size=13, bold=True, color=WHITE)
    ws3["A1"].fill = fill(ACCENT_TEAL)
    ws3["A1"].alignment = center()
    ws3.row_dimensions[1].height = 32

    # KPI cards
    total = len(df)
    avg_pct = df["attendance_percentage"].mean()
    below_75 = len(df[df["attendance_percentage"] < 75])
    perfect = len(df[df["attendance_percentage"] == 100])
    above_90 = len(df[df["attendance_percentage"] >= 90])

    kpis = [
        ("Total Students", total, MED_BLUE),
        ("Avg Attendance %", f"{avg_pct:.1f}%", ACCENT_TEAL),
        ("Below 75% (At Risk)", below_75, "C0392B"),
        ("Above 90%", above_90, GREEN),
        ("Perfect Attendance", perfect, "8E44AD"),
    ]

    for col_off, (label, val, color) in enumerate(kpis):
        col = col_off + 1
        ws3.column_dimensions[get_column_letter(col)].width = 20
        ws3.merge_cells(start_row=3, start_column=col, end_row=4, end_column=col)
        label_cell = ws3.cell(row=3, column=col, value=label)
        label_cell.font = Font(name="Arial", size=9, bold=True, color=WHITE)
        label_cell.fill = fill(color)
        label_cell.alignment = center()

        ws3.merge_cells(start_row=5, start_column=col, end_row=6, end_column=col)
        val_cell = ws3.cell(row=5, column=col, value=val)
        val_cell.font = Font(name="Arial", size=18, bold=True, color=color)
        val_cell.fill = fill(LIGHT_GREY)
        val_cell.alignment = center()
        ws3.row_dimensions[5].height = 32
        ws3.row_dimensions[3].height = 22

    # Distribution table
    ws3.row_dimensions[8].height = 20
    ws3.merge_cells("A8:F8")
    ws3["A8"] = "Attendance Distribution"
    ws3["A8"].font = Font(name="Arial", size=11, bold=True, color=DARK_BLUE)
    ws3["A8"].fill = fill(LIGHT_BLUE)
    ws3["A8"].alignment = center()

    dist_headers = ["Category", "Range", "Count", "% of Class", "Avg Present", "Avg Absent"]
    for c, h in enumerate(dist_headers, 1):
        cell = ws3.cell(row=9, column=c, value=h)
        cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
        cell.fill = fill(MED_BLUE)
        cell.alignment = center()
        cell.border = border

    categories = [
        ("Excellent",    90,  101, GREEN,    "E8F8F0"),
        ("Satisfactory", 75,   90, MED_BLUE, LIGHT_BLUE),
        ("At Risk",      60,   75, ORANGE,   "FEF3E2"),
        ("Critical",      0,   60, RED,      "FDEDEC"),
    ]

    for r_off, (cat, low, high, fc, bgc) in enumerate(categories):
        er = 10 + r_off
        mask = df["attendance_percentage"].between(low, high - 0.01)
        grp = df[mask]
        count = len(grp)
        pct_class = f"{count / total * 100:.1f}%" if total else "0%"
        avg_p = round(grp["total_present"].mean(), 1) if count else 0
        avg_a = round(grp["total_absent"].mean(), 1) if count else 0

        row_vals = [cat, f"{low}% – {high-1}%", count, pct_class, avg_p, avg_a]
        for c, v in enumerate(row_vals, 1):
            cell = ws3.cell(row=er, column=c, value=v)
            cell.font = Font(name="Arial", size=10, bold=(c == 1), color=fc)
            cell.fill = fill(bgc)
            cell.alignment = center()
            cell.border = border
        ws3.row_dimensions[er].height = 18

    # Top performers table
    ws3.row_dimensions[16].height = 20
    ws3.merge_cells("A16:F16")
    ws3["A16"] = "🏆  Top 5 Students by Attendance"
    ws3["A16"].font = Font(name="Arial", size=11, bold=True, color=DARK_BLUE)
    ws3["A16"].fill = fill(LIGHT_BLUE)
    ws3["A16"].alignment = center()

    top5_headers = ["Rank", "Student Name", "Total Classes", "Present", "Absent", "Attendance %"]
    for c, h in enumerate(top5_headers, 1):
        cell = ws3.cell(row=17, column=c, value=h)
        cell.font = Font(name="Arial", size=10, bold=True, color=WHITE)
        cell.fill = fill(GREEN)
        cell.alignment = center()
        cell.border = border

    top5 = df.nlargest(5, "attendance_percentage").reset_index(drop=True)
    medals = ["🥇", "🥈", "🥉", "4th", "5th"]
    for r_idx, row in top5.iterrows():
        er = 18 + r_idx
        for c, v in enumerate([medals[r_idx], row["student_name"], row["total_classes"],
                                row["total_present"], row["total_absent"],
                                f"{row['attendance_percentage']:.1f}%"], 1):
            cell = ws3.cell(row=er, column=c, value=v)
            cell.font = body_font(bold=(c == 2))
            cell.fill = fill(LIGHT_GREY if r_idx % 2 == 0 else WHITE)
            cell.alignment = center() if c != 2 else left()
            cell.border = border
        ws3.row_dimensions[er].height = 18

    # Save
    report_path = os.path.join(output_dir, "Attendance_Report.xlsx")
    wb.save(report_path)
    print(f"  ✓ Report Excel saved: {report_path}")
    return report_path



def main():
    print("  Smart Campus Management System")

    print("[1] Loading Excel file...")
    df_raw, date_row = load_attendance_sheet(EXCEL_FILE)
    print(f"    Found {len(df_raw)} raw student rows.")

    print("[2] Extracting & cleaning student data...")
    df_students = extract_student_summary(df_raw)
    print(f"    Processed {len(df_students)} students.")

    print("[3] Loading weekly report and merging student names...")
    df_weekly = load_weekly_report(EXCEL_FILE)
    df_student_report = build_student_report(df_students, df_weekly)
    print(f"    Built {len(df_student_report)} student report rows.")

    print("[4] Generating comprehensive analytics...")
    df_weekly_breakdown = extract_weekly_breakdown(df_raw)
    df_engagement = calculate_engagement_score(df_students)
    df_trends = calculate_attendance_trends(df_raw)
    df_percentiles = calculate_batch_percentiles(df_students)
    df_course_summary = generate_course_performance_summary(df_weekly)
    df_risk = generate_risk_assessment(df_students, df_engagement, df_trends)
    df_batch_analytics = generate_batch_analytics(df_students)
    df_demographic = generate_demographic_breakdown(df_students)
    
    analytics_dict = {
        "engagement_scores": df_engagement,
        "attendance_trends": df_trends,
        "batch_percentiles": df_percentiles,
        "course_analytics": df_course_summary,
        "risk_assessment": df_risk,
        "batch_analytics": df_batch_analytics,
        "demographic_breakdown": df_demographic,
        "weekly_breakdown": df_weekly_breakdown,
    }
    
    print(f"    ✓ Engagement scores: {len(df_engagement)} students")
    print(f"    ✓ Attendance trends: {len(df_trends)} students")
    print(f"    ✓ Course performance: {len(df_course_summary)} courses")
    print(f"    ✓ Risk assessment: {len(df_risk)} students")
    print(f"    ✓ Batch analytics: {len(df_batch_analytics)} batches")

    print("[5] Storing to SQLite database...")
    store_to_sqlite(df_students, df_student_report, DB_FILE, analytics_dict)

    print("[6] Exporting CSV files...")
    export_csvs(df_students, OUTPUT_DIR)
    export_student_report_csv(df_student_report, OUTPUT_DIR)
    export_analytics_csvs(analytics_dict, OUTPUT_DIR)

    print("[7] Generating Excel report...")
    report_path = generate_report_excel(df_students, OUTPUT_DIR)

    print("\n" + "="*60)
    print("  COMPREHENSIVE ANALYSIS SUMMARY")
    print("="*60)
    print(f"  Total students              : {len(df_students)}")
    print(f"  Average attendance          : {df_students['attendance_percentage'].mean():.1f}%")
    print(f"  Below 75% (at-risk)         : {len(df_students[df_students['attendance_percentage'] < 75])}")
    print(f"  Perfect attendance (100%)   : {len(df_students[df_students['attendance_percentage'] == 100])}")
    print(f"\n  Engagement Tiers:")
    for tier in ["Highly Engaged", "Engaged", "Moderately Engaged", "Low Engagement", "Disengaged"]:
        count = (df_engagement["engagement_tier"] == tier).sum()
        if count > 0:
            print(f"    {tier}: {count} students")
    print(f"\n  Risk Levels:")
    for level in ["Critical", "High", "Medium", "Low"]:
        count = (df_risk["risk_level"] == level).sum()
        print(f"    {level}: {count} students")
    print(f"\n  Course Performance:")
    for _, course in df_course_summary.iterrows():
        print(f"    {course['course_name']}: {course['avg_attendance_pct']:.1f}% avg")
    print(f"\n  Database  : {DB_FILE}")
    print(f"  Outputs   : {OUTPUT_DIR}/")
    print("="*60)

    return df_students


if __name__ == "__main__":
    main()